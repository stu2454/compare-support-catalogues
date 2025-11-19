"""
XLSX to JSON Converter for NDIS Support Catalogue

Handles:
- Sheet detection and listing
- User-driven sheet mapping
- Robust conversion with normalization
- Handling of different column naming conventions
"""

import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import openpyxl


def get_sheet_info(file_path: str) -> List[Dict[str, any]]:
    """
    Get information about all sheets in the workbook.
    
    Returns list of dicts with: {name, row_count, col_count}
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets_info = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        sheets_info.append({
            'name': sheet_name,
            'row_count': max_row,
            'col_count': max_col
        })
    
    wb.close()
    return sheets_info


def preview_sheet(file_path: str, sheet_name: str, n_rows: int = 10) -> pd.DataFrame:
    """
    Preview the first n rows of a specific sheet.
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=n_rows)
        return df
    except Exception as e:
        return pd.DataFrame({'Error': [f'Could not preview sheet: {str(e)}']})


def normalize_value(value, field_type: str = 'string'):
    """
    Normalize a cell value based on expected type.
    
    field_type can be: 'string', 'boolean', 'date', 'float', 'int'
    """
    # Handle None/NaN
    if pd.isna(value) or value is None or value == '':
        return None
    
    if field_type == 'boolean':
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            value_lower = value.strip().lower()
            if value_lower in ['yes', 'y', 'true', '1']:
                return True
            elif value_lower in ['no', 'n', 'false', '0']:
                return False
        return None
    
    elif field_type == 'date':
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
        if isinstance(value, str):
            # Try to parse common date formats
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                try:
                    dt = datetime.strptime(value.strip(), fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        return None
    
    elif field_type == 'float':
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    
    elif field_type == 'int':
        try:
            return int(value)
        except (ValueError, TypeError):
            return None
    
    else:  # string
        return str(value).strip() if value else None


def detect_column_mapping(df: pd.DataFrame) -> Dict[str, str]:
    """
    Detect and map actual column names to standardized field names.
    
    Handles variations like:
    - 'Item Number' vs 'Support Item Number' vs 'item_number'
    - 'Description' vs 'Support Name' vs 'Name'
    - State columns: 'NSW', 'New South Wales', 'NSW Price Limit', etc.
    """
    columns_lower = {col.lower().strip(): col for col in df.columns}
    
    mapping = {}
    
    # Item Number
    for pattern in ['item number', 'item_number', 'support item number', 'support_item_number', 'item no']:
        if pattern in columns_lower:
            mapping['item_number'] = columns_lower[pattern]
            break
    
    # Support Name / Description
    for pattern in ['support name', 'support_name', 'description', 'name', 'support item name']:
        if pattern in columns_lower:
            mapping['support_name'] = columns_lower[pattern]
            break
    
    # Registration Group
    for pattern in ['registration group', 'registration_group', 'reg group', 'group']:
        if pattern in columns_lower:
            mapping['registration_group'] = columns_lower[pattern]
            break
    
    # Category
    for pattern in ['category', 'support category', 'support_category']:
        if pattern in columns_lower:
            mapping['category'] = columns_lower[pattern]
            break
    
    # Unit of Measure
    for pattern in ['unit', 'unit of measure', 'uom', 'unit_of_measure']:
        if pattern in columns_lower:
            mapping['unit'] = columns_lower[pattern]
            break
    
    # Claim Type
    for pattern in ['claim type', 'claim_type', 'claiming']:
        if pattern in columns_lower:
            mapping['claim_type'] = columns_lower[pattern]
            break
    
    # State price limits - Australian states
    states = {
        'nsw': ['nsw', 'new south wales'],
        'vic': ['vic', 'victoria'],
        'qld': ['qld', 'queensland'],
        'sa': ['sa', 'south australia'],
        'wa': ['wa', 'western australia'],
        'tas': ['tas', 'tasmania'],
        'nt': ['nt', 'northern territory'],
        'act': ['act', 'australian capital territory']
    }
    
    for state_key, patterns in states.items():
        for pattern in patterns:
            # Look for exact match or "STATE Price" or "Price STATE"
            for col_lower, col_actual in columns_lower.items():
                if (pattern == col_lower or 
                    f'{pattern} price' in col_lower or 
                    f'price {pattern}' in col_lower or
                    f'{pattern}_price' in col_lower or
                    f'price_{pattern}' in col_lower or
                    f'{pattern} limit' in col_lower or
                    f'price limit {pattern}' in col_lower):
                    mapping[f'price_limit_{state_key}'] = col_actual
                    break
            if f'price_limit_{state_key}' in mapping:
                break
    
    # Dates
    for pattern in ['effective from', 'effective_from', 'start date', 'from date']:
        if pattern in columns_lower:
            mapping['effective_from'] = columns_lower[pattern]
            break
    
    for pattern in ['effective to', 'effective_to', 'end date', 'to date']:
        if pattern in columns_lower:
            mapping['effective_to'] = columns_lower[pattern]
            break
    
    # Notes
    for pattern in ['notes', 'note', 'comments', 'conditions']:
        if pattern in columns_lower:
            mapping['notes'] = columns_lower[pattern]
            break
    
    return mapping


def convert_sheet_to_items(df: pd.DataFrame, sheet_name: str) -> List[Dict]:
    """
    Convert a DataFrame (from one sheet) into a list of normalized item dicts.
    """
    if df.empty:
        return []
    
    # Detect column mapping
    col_map = detect_column_mapping(df)
    
    items = []
    
    for idx, row in df.iterrows():
        # Must have an item_number to be valid
        if 'item_number' not in col_map:
            continue
        
        item_number = normalize_value(row[col_map['item_number']], 'string')
        if not item_number:
            continue  # Skip rows without item number
        
        item = {
            'item_number': item_number,
            'support_name': normalize_value(row[col_map.get('support_name', '')], 'string') if 'support_name' in col_map else None,
            'registration_group': normalize_value(row[col_map.get('registration_group', '')], 'string') if 'registration_group' in col_map else None,
            'category': normalize_value(row[col_map.get('category', '')], 'string') if 'category' in col_map else None,
            'unit': normalize_value(row[col_map.get('unit', '')], 'string') if 'unit' in col_map else None,
            'claim_type': normalize_value(row[col_map.get('claim_type', '')], 'string') if 'claim_type' in col_map else None,
            'effective_from': normalize_value(row[col_map.get('effective_from', '')], 'date') if 'effective_from' in col_map else None,
            'effective_to': normalize_value(row[col_map.get('effective_to', '')], 'date') if 'effective_to' in col_map else None,
            'notes': normalize_value(row[col_map.get('notes', '')], 'string') if 'notes' in col_map else None,
            'raw_row_index': int(idx) + 2  # +2 because Excel is 1-indexed and has header row
        }
        
        # Add state price limits
        for state in ['nsw', 'vic', 'qld', 'sa', 'wa', 'tas', 'nt', 'act']:
            col_key = f'price_limit_{state}'
            if col_key in col_map:
                item[col_key] = normalize_value(row[col_map[col_key]], 'float')
            else:
                item[col_key] = None
        
        items.append(item)
    
    return items


def convert_catalogue_to_json(
    file_path: str,
    current_sheet_name: str,
    legacy_sheet_name: Optional[str] = None,
    source_filename: Optional[str] = None
) -> Dict:
    """
    Convert a full Support Catalogue workbook to structured JSON.
    
    Args:
        file_path: Path to the Excel file
        current_sheet_name: Name of the sheet containing current items
        legacy_sheet_name: Name of the sheet containing legacy items (or None)
        source_filename: Original filename for metadata
    
    Returns:
        Dict with structure:
        {
            "metadata": {...},
            "current_items": [...],
            "legacy_items": [...]
        }
    """
    # Read current sheet
    df_current = pd.read_excel(file_path, sheet_name=current_sheet_name)
    current_items = convert_sheet_to_items(df_current, current_sheet_name)
    
    # Read legacy sheet if provided
    legacy_items = []
    legacy_meta = None
    if legacy_sheet_name:
        df_legacy = pd.read_excel(file_path, sheet_name=legacy_sheet_name)
        legacy_items = convert_sheet_to_items(df_legacy, legacy_sheet_name)
        legacy_meta = {
            'sheet': legacy_sheet_name,
            'rows': len(df_legacy)
        }
    
    # Build metadata
    metadata = {
        'source_filename': source_filename or file_path,
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'current': {
            'sheet': current_sheet_name,
            'rows': len(df_current)
        }
    }
    
    if legacy_meta:
        metadata['legacy'] = legacy_meta
    
    return {
        'metadata': metadata,
        'current_items': current_items,
        'legacy_items': legacy_items
    }
